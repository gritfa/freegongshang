"""Excel数据读取模块"""
from openpyxl import load_workbook
from loguru import logger
from typing import List, Dict, Optional


def read_enterprise_data(filepath: str) -> List[Dict[str, str]]:
    """读取企业基本信息Excel
    
    Excel字段映射（根据截图）:
    A列: 企业名称
    B列: 注册号/统一社会信用代码
    C列: 地址
    D列: 法定代表人
    E列: 身份证
    
    Returns:
        企业信息字典列表
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    
    enterprises = []
    headers = []
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx == 1:
            # 读取表头
            headers = [str(cell).strip() if cell else "" for cell in row]
            logger.info(f"Excel表头: {headers}")
            continue
        
        if not row[0]:  # 跳过空行
            continue
        
        enterprise = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                enterprise[headers[col_idx]] = str(value).strip() if value else ""
        
        enterprises.append(enterprise)
    
    wb.close()
    logger.info(f"读取到 {len(enterprises)} 家企业数据")
    return enterprises


def read_annual_report_data(filepath: str) -> Dict[str, Dict[str, str]]:
    """读取年报数据Excel
    
    以注册号为key，方便按企业查找
    
    Returns:
        {注册号: {字段名: 值}} 的字典
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    
    report_data = {}
    headers = []
    reg_col_idx = None  # 注册号所在列的索引
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx == 1:
            headers = [str(cell).strip() if cell else "" for cell in row]
            logger.info(f"年报数据表头: {headers}")
            # 找到注册号列
            for i, h in enumerate(headers):
                if "注册号" in h or "信用代码" in h or "社会信用" in h:
                    reg_col_idx = i
                    break
            if reg_col_idx is None:
                logger.warning("未找到注册号列，默认使用第2列(B列)")
                reg_col_idx = 1
            continue
        
        if not row[0]:
            continue
        
        reg_no = str(row[reg_col_idx]).strip() if row[reg_col_idx] else ""
        if not reg_no:
            continue
        
        data = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                data[headers[col_idx]] = str(value).strip() if value else ""
        
        report_data[reg_no] = data
    
    wb.close()
    logger.info(f"读取到 {len(report_data)} 家企业的年报数据")
    return report_data
